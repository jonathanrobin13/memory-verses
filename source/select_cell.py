from openpyxl import Workbook
from files import file
from length_formatter import length_formatter

INPUT_EXCEL_FILE = "Excel_Format.xlsx"
OUTPUT_EXCEL_FILE = "Selected_Verses.xlsx"


wb = file(INPUT_EXCEL_FILE, True)
sheet = wb.active

# selected_verses_xlsx = Workbook()

# new_sheet = selected_verses_xlsx.active
# new_sheet.title = "Verses"

selected_verses_xlsx = file(OUTPUT_EXCEL_FILE, True)
new_sheet = selected_verses_xlsx['Verses']

new_sheet.delete_rows(2, new_sheet.max_row)

print("Type all the references  that you want. Split each reference by a comma.")
print("For example, Proverbs 3:16,Genesis 1:1,John 3:16")
references = input()

verses_to_get = references.split(",")

open_row = 2

for reference in verses_to_get:
    while True:
        found = False

        for row in sheet.iter_rows(min_row=2):
            if reference in row[1].value:
                verse = row[2].value
                new_sheet.cell(open_row, 3).value = verse
                found = True
                break

        if found:
            break
        else:
            reference = input(f"{reference} not found, please type it again\n")

    new_sheet.cell(open_row, 2).value = reference
    new_sheet.cell(open_row, 1).value = open_row - 1

    open_row += 1

selected_verses_xlsx.save(file(OUTPUT_EXCEL_FILE, False))

length_formatter()
