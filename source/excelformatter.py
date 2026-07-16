import openpyxl as xl
import fitz
from files import file


# Open excel sheet

excel_format_xlsx = file("Excel_Format.xlsx", True)
ws = excel_format_xlsx["Verses"]


# Open PDF
doc = fitz.open(file("Bible_Verses.pdf", False))


for page in doc:

    blocks = page.get_text("dict")["blocks"]

    for block in blocks:

        if "lines" not in block:
            continue

        for line in block["lines"]:

            for span in line["spans"]:

                if span["text"] == "5. Exodus":
                    open_row += 1

                    ws.cell(open_row, 1).value = "5"
                    ws.cell(open_row, 2).value = "Exodus 34:7"
                    ws.cell(open_row, 3).value = "Keeping mercy for thousands, forgiving iniquity and transgression and sin, by no means clearing the guilty, visiting the iniquity of the fathers upon the children and the children’s children to the third and the fourth generation."

                # If the font in bold, then that means it is a verse heading
                elif span["font"] == "TimesNewRomanPS-BoldMT":

                    complete_verse = ""

                    # sometimes there are empty spans and headings that don't have verses and there are headings, so skip them
                    if span["text"] == ' ' or "School" in span["text"] or span["text"] == "300 Bible Verses":
                        continue

                    # If the text is bolded, then that means that there was a verse completed
                    try:
                        for verse in verses_list:
                            complete_verse += verse

                        cell_verse.value = complete_verse

                        # Put the verses in the excel sheet

                    except NameError:
                        pass

                    # splits verse number and reference
                    verse_heading = span["text"].split(sep=".", maxsplit=1)
                    verse_number = int(verse_heading[0])
                    reference = verse_heading[1]

                    open_row = verse_number + 1

                    # create cell
                    # add + 1 to verse_number so I can put heading for verse #, ref, and verse
                    verse_num_cell = ws.cell(open_row, 1)
                    reference_cell = ws.cell(open_row, 2)

                    # Put in the values
                    verse_num_cell.value = verse_number
                    reference_cell.value = reference

                    verses_list = []

                elif span["font"] == "TimesNewRomanPSMT" and span["text"] != ' ':

                    cell_verse = ws.cell(open_row, 3)
                    verses_list.append(span["text"])


excel_format_xlsx.save(file("Excel_Format.xlsx", False))
