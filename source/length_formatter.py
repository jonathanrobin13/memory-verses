import openpyxl
from files import file


def length_formatter():

    INPUT_EXCEL_FILE = "Excel_Format.xlsx"
    OUTPUT_EXCEL_FILE = "verses_sorted.xlsx"

    esv_verses_xlsx = file(INPUT_EXCEL_FILE, True)
    old_ws = esv_verses_xlsx.active

    verses_sorted_xlsx = file(OUTPUT_EXCEL_FILE, True)
    new_ws = verses_sorted_xlsx["Sorted Verses"]

    new_ws.delete_rows(2, new_ws.max_row)

    rows = []

    # Skip the blank first row
    for row in old_ws.iter_rows(min_row=2, values_only=True):

        reference = row[1]   # Column B
        verse = row[2]       # Column C
        verse = verse.replace('\n', ' ')

        # Count every character
        length = len(verse) if verse else 0

        rows.append((length, reference, verse))

    # Sort by verse length
    rows.sort(key=lambda x: x[0])

    # Write the sorted data
    for i, (_, reference, verse) in enumerate(rows, start=2):
        new_ws.cell(row=i, column=1).value = i - 1
        new_ws.cell(row=i, column=2).value = reference
        new_ws.cell(row=i, column=3).value = verse

    new_ws.cell(1, 1).value = "Verse Number"
    new_ws.cell(1, 2).value = "Reference"
    new_ws.cell(1, 3).value = "Verse"

    # Save the new workbook
    verses_sorted_xlsx.save(file(OUTPUT_EXCEL_FILE, False))


if __name__ == "main":
    length_formatter()
