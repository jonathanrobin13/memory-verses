import openpyxl as xl

from pathlib import Path

from dotenv import load_dotenv
import os
import requests

from tqdm import tqdm

import time


project_folder = Path(__file__).parent.parent
data_folder = project_folder / "data"

excel_format_file = data_folder / "Excel_Format.xlsx"
wb = xl.load_workbook(excel_format_file)
ws = wb["Verses"]

esv_verses_file = data_folder / "ESV_Verses.xlsx"
wb_new = xl.load_workbook(esv_verses_file)
ws_new = wb_new['Verses']

load_dotenv()

ESV_API_KEY = os.getenv("ESV_API_KEY")

url = "https://api.esv.org/v3/passage/text/"

headers = {
    "Authorization": f"Token {ESV_API_KEY}"
}

params = {
    "include-passage-references": False,  # make false
    "include-verse-numbers": False,
    "include-first-verse-numbers": False,
    "include-footnotes": False,
    "include-footnote-body": False,
    "include-headings": False,
    "include-short-copyright": False,
    "indent-poetry": False,
    "include-selahs": False
}


references = []

session = requests.Session()

for row in tqdm(range(2, 302)):
    reference_cell = ws.cell(row, 2)
    reference = reference_cell.value

    params["q"] = reference

    response = session.get(
        url,
        headers=headers,
        params=params
    )

    verse = response.json()

    verse_cell = ws_new.cell(row, 3)

    try:
        verse_cell.value = verse['passages'][0].strip()
    except KeyError:
        print(verse)
        print(row)
        time.sleep(60)

session.close()


wb_new.save(r"C:\Users\robin\memory-verse\data\ESV Verses.xlsx")


#     for index in range(0, len(references)):
#         if index == 0:
#             params["q"] = references[index]
#             continue

#         params["q"] = params["q"] + ";" + references[index]
