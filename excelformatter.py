import openpyxl as xl
import fitz

# Open execl sheet
wb = xl.load_workbook("Memory_Verses.xlsx")
old_verses = wb["Old_Verses"]
new_verses = wb["New Verses"]

# Open PDF
doc = fitz.open("300_bible_verses_final_2026.pdf")

i = 0

# for page in doc:

#     blocks = page.get_text("dict")["blocks"]

blocks = doc[0].get_text("dict")["blocks"]

for block in blocks:

    if "lines" not in block:
        continue

    for line in block["lines"]:

        for span in line["spans"]:

            if span["text"] == "5. Exodus 34:7":
                open_row += 1

                old_verses.cell(open_row, 1).value = "5"
                old_verses.cell(open_row, 2).value = "Exodus 34:7"
                old_verses.cell(open_row, 3).value = "Keeping mercy for thousands, forgiving iniquity and transgression and sin, by no means clearing the guilty, visiting the iniquity of the fathers upon the children and the children’s children to the third and the fourth generation."

            # If the font in bold, then that means it is a verse heading
            elif span["font"] == "TimesNewRomanPS-BoldMT":

                # sometimes there are empty spans and headings that don't have verses, so skip them
                if span["text"] == ' ' or span["text"] == '' or "PCP" in span["text"] or span["text"] == "300 Bible Verses":
                    continue

                # splits verse number and reference
                verse_heading = span["text"].split(sep=".", maxsplit=1)
                verse_number = int(verse_heading[0])
                reference = verse_heading[1]

                open_row = verse_number + 1

                # create cell
                # add + 1 to verse_number so I can put heading for verse #, ref, and verse
                verse_num_cell = old_verses.cell(open_row, 1)
                reference_cell = old_verses.cell(open_row, 2)

                # Put in the values
                verse_num_cell.value = verse_number
                reference_cell.value = reference

            elif span["font"] == "TimesNewRomanPSMT" and span["text"] != " ":
                cell_verse = old_verses.cell(open_row, 3)

                cell_verse.value = span["text"]


wb.save("Memory_Verses.xlsx")
